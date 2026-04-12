from __future__ import annotations

import re
from io import BytesIO
from datetime import UTC, date, datetime, time
from typing import TypedDict, cast
from uuid import UUID

from openpyxl import load_workbook  # type: ignore[import-untyped]
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.journal import ActivityEntry, ActivityStatus, ActivityType
from app.models.user import User
from app.repositories.journal import JournalRepository
from app.schemas.journal import (
    ActivityEntryCreateRequest,
    ActivityEntryUpdateRequest,
    ActivityStatus as JournalSchemaStatus,
    ActivityType as JournalSchemaType,
    BulkJournalImportRequest,
    BulkJournalImportPreviewItem,
)

SECTION_DATE_RE = re.compile(
    r"^(?P<label>Выполненные задачи|Взята в работу|Выполнено|В работе)\s*"
    r"(?P<date>\d{2}\.\d{2}\.\d{2,4})?\s*:?\s*$"
)
MARKDOWN_LINK_RE = re.compile(r"\[(?P<title>[^\]]+)\]\((?P<url>[^)]+)\)")
TICKET_NUMBER_RE = re.compile(
    r"\b(?:SR|Задача)?\s*(?P<number>\d{5,})\b",
    re.IGNORECASE,
)
_MAX_ACTIVITY_TITLE_LENGTH = 255
_EXCEL_IMPORT_REQUIRED_HEADERS = {
    "Номер": "ticket_number",
    "Услуга": "service",
    "Фактическое разрешение": "resolved_at",
}


class ParsedTextImportItem(TypedDict):
    """Нормализованная запись текстового импорта журнала."""

    work_date: date
    activity_type: JournalSchemaType
    status: JournalSchemaStatus
    title: str
    service: str | None
    ticket_number: str | None
    task_url: str | None


class ParsedExcelImportItem(TypedDict):
    """Нормализованная запись Excel-импорта журнала."""

    work_date: date
    activity_type: JournalSchemaType
    status: JournalSchemaStatus
    title: str
    service: str | None
    ticket_number: str
    resolved_at: datetime


async def list_activity_entries_for_date(
    session: AsyncSession,
    user_id: str,
    work_date: date,
) -> list[ActivityEntry]:
    """Возвращает записи пользователя за конкретную рабочую дату.

    Почему фильтрация идёт именно по user_id + work_date:
    - пользователь должен видеть только свои записи;
    - одна и та же дата может дополняться позже;
    - отчёт за день строится по work_date, а не по created_at.
    """
    repo = JournalRepository(session)
    return await repo.list_for_date(user_id=user_id, work_date=work_date)


async def create_activity_entry(
    session: AsyncSession,
    user: User,
    payload: ActivityEntryCreateRequest,
) -> ActivityEntry:
    """Создаёт новую запись журнала для текущего пользователя.

    Логика работы с датами:
    - work_date отвечает за день, к которому запись относится в отчётах;
    - ended_date отвечает за реальную дату закрытия задачи;
    - если ended_date не передана, считаем, что запись закрыта в work_date.
    """
    repo = JournalRepository(session)

    effective_ended_date = payload.ended_date or payload.work_date
    if effective_ended_date < payload.work_date:
        raise ValueError("Дата окончания не может быть раньше рабочей даты")

    started_at_value = payload.started_at
    if started_at_value is None:
        last_finished = await repo.get_last_finished_time_for_date(
            user_id=user.id,
            work_date=payload.work_date,
        )
        if last_finished is not None:
            started_at_value = last_finished.timetz().replace(tzinfo=None)

    started_at_dt = _combine_work_date_and_time(payload.work_date, started_at_value)
    ended_at_dt = _combine_work_date_and_time(effective_ended_date, payload.ended_at)
    if (
        started_at_dt is not None
        and ended_at_dt is not None
        and ended_at_dt < started_at_dt
    ):
        raise ValueError("Время окончания не может быть раньше времени начала")

    activity_entry = ActivityEntry(
        user_id=user.id,
        work_date=payload.work_date,
        activity_type=payload.activity_type,
        status=payload.status,
        title=payload.title.strip(),
        description=payload.description.strip() if payload.description else None,
        resolution=payload.resolution.strip() if payload.resolution else None,
        contact=payload.contact.strip() if payload.contact else None,
        service=payload.service.strip() if payload.service else None,
        external_ref=payload.ticket_number.strip() if payload.ticket_number else None,
        ticket_number=payload.ticket_number.strip() if payload.ticket_number else None,
        task_url=payload.task_url.strip() if payload.task_url else None,
        started_at=started_at_dt,
        finished_at=ended_at_dt,
    )

    return await repo.save(activity_entry)


async def get_activity_entry_by_id(
    session: AsyncSession,
    user_id: str,
    entry_id: str,
) -> ActivityEntry | None:
    """Возвращает запись по ID в рамках пользователя-владельца."""
    repo = JournalRepository(session)
    return await repo.get_by_id(user_id=user_id, entry_id=entry_id)


async def update_activity_entry(
    session: AsyncSession,
    entry: ActivityEntry,
    payload: ActivityEntryUpdateRequest,
) -> ActivityEntry:
    """Обновляет запись журнала по переданным полям.

    При обновлении важно сначала вычислить итоговые значения даты и времени,
    а уже потом проверять ограничения. Иначе можно пропустить некорректную
    комбинацию полей, если часть значений пришла в запросе, а часть осталась
    в существующей записи.
    """
    repo = JournalRepository(session)

    original_work_date = entry.work_date
    original_started_at = entry.started_at
    original_finished_at = entry.finished_at

    effective_work_date = (
        payload.work_date if payload.work_date is not None else original_work_date
    )
    effective_started_time = _resolve_updated_time(
        field_name="started_at",
        payload=payload,
        current_value=original_started_at,
    )
    effective_ended_time = _resolve_updated_time(
        field_name="ended_at",
        payload=payload,
        current_value=original_finished_at,
    )
    effective_ended_date = _resolve_updated_end_date(
        payload=payload,
        original_work_date=original_work_date,
        original_finished_at=original_finished_at,
        effective_work_date=effective_work_date,
    )

    if effective_ended_date is not None and effective_ended_date < effective_work_date:
        raise ValueError("Дата окончания не может быть раньше рабочей даты")

    effective_started_dt = _combine_work_date_and_time(
        effective_work_date, effective_started_time
    )
    effective_finished_dt = _combine_work_date_and_time(
        (
            effective_ended_date
            if effective_ended_date is not None
            else effective_work_date
        ),
        effective_ended_time,
    )
    if (
        effective_started_dt is not None
        and effective_finished_dt is not None
        and effective_finished_dt < effective_started_dt
    ):
        raise ValueError("Время окончания не может быть раньше времени начала")

    if "work_date" in payload.model_fields_set and payload.work_date is not None:
        entry.work_date = payload.work_date
    if (
        "activity_type" in payload.model_fields_set
        and payload.activity_type is not None
    ):
        entry.activity_type = payload.activity_type
    if "status" in payload.model_fields_set and payload.status is not None:
        entry.status = payload.status
    if "title" in payload.model_fields_set and payload.title is not None:
        entry.title = payload.title.strip()
    if "description" in payload.model_fields_set:
        entry.description = _normalize_optional_text(payload.description)
    if "resolution" in payload.model_fields_set:
        entry.resolution = _normalize_optional_text(payload.resolution)
    if "contact" in payload.model_fields_set:
        entry.contact = _normalize_optional_text(payload.contact)
    if "service" in payload.model_fields_set:
        entry.service = _normalize_optional_text(payload.service)
    if "ticket_number" in payload.model_fields_set:
        normalized_ticket = _normalize_optional_text(payload.ticket_number)
        entry.ticket_number = normalized_ticket
        entry.external_ref = normalized_ticket
    if "task_url" in payload.model_fields_set:
        entry.task_url = _normalize_optional_text(payload.task_url)

    if "started_at" in payload.model_fields_set:
        entry.started_at = effective_started_dt
    elif "work_date" in payload.model_fields_set and original_started_at is not None:
        entry.started_at = effective_started_dt

    if "ended_at" in payload.model_fields_set:
        entry.finished_at = (
            effective_finished_dt if payload.ended_at is not None else None
        )
    elif "ended_date" in payload.model_fields_set:
        if payload.ended_date is not None:
            entry.finished_at = effective_finished_dt
    elif "work_date" in payload.model_fields_set and original_finished_at is not None:
        entry.finished_at = effective_finished_dt

    return await repo.update(entry)


async def delete_activity_entry(session: AsyncSession, entry: ActivityEntry) -> None:
    """Удаляет запись журнала."""
    repo = JournalRepository(session)
    await repo.delete(entry)


async def delete_activity_entries_for_date(
    session: AsyncSession,
    user_id: str,
    work_date: date,
) -> int:
    """Удаляет все записи пользователя за выбранную рабочую дату.

    Это действие нужно для сценария полной очистки конкретного дня:
    пользователь может ошибочно импортировать или создать много записей,
    и тогда удобнее удалить весь день целиком, чем очищать его вручную по одной.
    """
    repo = JournalRepository(session)
    return await repo.delete_for_date(user_id=user_id, work_date=work_date)


async def delete_all_activity_entries(
    session: AsyncSession,
    user_id: str,
) -> int:
    """Удаляет все записи журнала текущего пользователя.

    Это максимально опасное действие, поэтому само бизнес-правило остаётся простым:
    удаляем только собственные записи пользователя и никогда не затрагиваем чужой журнал.
    """
    repo = JournalRepository(session)
    return await repo.delete_all(user_id=user_id)


async def delete_selected_activity_entries(
    session: AsyncSession,
    user_id: str,
    entry_ids: list[str],
) -> int:
    """Удаляет только выбранные записи журнала текущего пользователя.

    Важное ограничение безопасности:
    - удаляем только записи, которые принадлежат текущему пользователю;
    - неизвестные или чужие id просто не затрагиваются.
    """
    repo = JournalRepository(session)
    return await repo.delete_selected(user_id=user_id, entry_ids=entry_ids)


async def delete_duplicate_activity_entries_for_date(
    session: AsyncSession,
    user_id: str,
    work_date: date,
) -> tuple[int, list[str]]:
    """Удаляет дубли записей журнала за рабочую дату.

    Дублем считаем повторную запись с тем же `ticket_number` в рамках
    той же `work_date` и того же пользователя.
    Сохраняем самую раннюю запись, потому что она обычно является
    первичной фиксацией работы, а остальные появляются из повторного импорта.
    """
    repo = JournalRepository(session)
    entries = await repo.list_with_ticket_for_date(user_id=user_id, work_date=work_date)

    removed_count = 0
    duplicate_ticket_numbers: list[str] = []
    seen_ticket_numbers: set[str] = set()

    for entry in entries:
        normalized_ticket_number = (entry.ticket_number or "").strip()
        if not normalized_ticket_number:
            continue

        if normalized_ticket_number not in seen_ticket_numbers:
            seen_ticket_numbers.add(normalized_ticket_number)
            continue

        duplicate_ticket_numbers.append(normalized_ticket_number)
        await session.delete(entry)
        removed_count += 1

    if removed_count == 0:
        return 0, []

    await session.commit()
    return removed_count, sorted(set(duplicate_ticket_numbers))


async def import_activity_entries_from_text(
    session: AsyncSession,
    user: User,
    payload: BulkJournalImportRequest,
) -> tuple[list[ActivityEntry], list[str]]:
    """Импортирует журнал из текста с секциями по датам.

    Поддерживаем простой формат:
    - заголовок секции с датой, например "Выполненные задачи 10.04.26";
    - строки с SR/задачами под этим заголовком;
    - секцию "Взята в работу" для задач в процессе.
    """
    repo = JournalRepository(session)
    parsed_items, warnings = _parse_bulk_import_text(payload)

    created_entries: list[ActivityEntry] = []
    for item in parsed_items:
        created_entries.append(
            ActivityEntry(
                user_id=user.id,
                work_date=item["work_date"],
                activity_type=item["activity_type"],
                status=item["status"],
                title=item["title"],
                service=item["service"],
                description=None,
                resolution=None,
                contact=None,
                external_ref=item["ticket_number"],
                ticket_number=item["ticket_number"],
                task_url=item["task_url"],
                started_at=None,
                finished_at=None,
            )
        )

    if not created_entries:
        raise ValueError("Не удалось распознать записи для импорта")

    return await repo.save_all(created_entries), warnings


async def import_activity_entries_from_excel_workbook(
    session: AsyncSession,
    user: User,
    workbook_bytes: bytes,
) -> tuple[list[ActivityEntry], list[str]]:
    """Импортирует записи журнала из Excel-файла.

    Бизнес-правило для этого сценария:
    - каждая строка считается завершённой заявкой;
    - рабочая дата записи равна дате фактического разрешения;
    - номер заявки и услуга переносятся в журнал как основные атрибуты;
    - дубликаты по связке `ticket_number + work_date` не создаём повторно.
    """
    repo = JournalRepository(session)
    parsed_items, warnings = parse_excel_workbook_preview(workbook_bytes)
    if not parsed_items:
        raise ValueError("Не удалось найти корректные строки в Excel-файле")

    ticket_date_pairs = {
        (str(item["ticket_number"]), item["work_date"]) for item in parsed_items
    }
    existing_pairs = await repo.get_existing_ticket_pairs(
        user_id=user.id,
        ticket_date_pairs=ticket_date_pairs,
    )

    created_entries: list[ActivityEntry] = []
    duplicate_count = 0
    for item in parsed_items:
        pair = (item["ticket_number"], item["work_date"])
        if pair in existing_pairs:
            duplicate_count += 1
            continue

        resolved_at = item["resolved_at"]
        created_entries.append(
            ActivityEntry(
                user_id=user.id,
                work_date=item["work_date"],
                activity_type=item["activity_type"],
                status=item["status"],
                title=item["title"],
                service=item["service"],
                description=None,
                resolution=None,
                contact=None,
                external_ref=item["ticket_number"],
                ticket_number=item["ticket_number"],
                task_url=None,
                started_at=None,
                finished_at=_combine_work_date_and_time(
                    item["work_date"], resolved_at.time()
                ),
            )
        )
        existing_pairs.add(pair)

    if duplicate_count:
        warnings.append(
            "Часть строк пропущена как дубликаты по номеру заявки и дате решения: "
            f"{duplicate_count}"
        )

    if not created_entries:
        raise ValueError(
            "Все строки из Excel-файла уже есть в журнале или не содержат корректных данных"
        )

    return await repo.save_all(created_entries), warnings


def preview_activity_entries_from_excel_workbook(
    workbook_bytes: bytes,
) -> tuple[list[BulkJournalImportPreviewItem], list[str]]:
    """Возвращает предпросмотр Excel-импорта без сохранения в базу."""
    parsed_items, warnings = parse_excel_workbook_preview(workbook_bytes)
    return [
        BulkJournalImportPreviewItem(
            work_date=item["work_date"],
            activity_type=item["activity_type"],
            status=item["status"],
            title=item["title"],
            service=item["service"],
            ticket_number=item["ticket_number"],
            task_url=None,
        )
        for item in parsed_items
    ], warnings


def preview_activity_entries_from_text(
    payload: BulkJournalImportRequest,
) -> tuple[list[BulkJournalImportPreviewItem], list[str]]:
    """Возвращает нормализованный предпросмотр без сохранения в базу."""
    parsed_items, warnings = _parse_bulk_import_text(payload)
    return [
        BulkJournalImportPreviewItem(
            work_date=item["work_date"],
            activity_type=item["activity_type"],
            status=item["status"],
            title=item["title"],
            service=item["service"],
            ticket_number=item["ticket_number"],
            task_url=item["task_url"],
        )
        for item in parsed_items
    ], warnings


def parse_excel_workbook_preview(
    workbook_bytes: bytes,
) -> tuple[list[ParsedExcelImportItem], list[str]]:
    """Разбирает Excel-файл в нормализованные записи журнала.

    Мы явно читаем только нужные для журнала поля:
    - `Номер`;
    - `Услуга`;
    - `Фактическое разрешение`.

    Остальные столбцы намеренно не тащим в журнал,
    чтобы не смешивать служебные детали выгрузки с краткой операционной записью.
    """
    workbook = _load_excel_workbook(workbook_bytes)
    worksheet = workbook.worksheets[0]
    header_map = _build_excel_header_map(worksheet)

    items: list[ParsedExcelImportItem] = []
    warnings: list[str] = []

    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        ticket_number = _normalize_excel_ticket_number(row[header_map["ticket_number"]])
        service = _normalize_optional_text(row[header_map["service"]])
        resolved_at = _normalize_excel_resolved_at(row[header_map["resolved_at"]])

        if not ticket_number and not service and resolved_at is None:
            continue
        if not ticket_number:
            warnings.append(f"Строка {row_index} пропущена: не найден номер заявки")
            continue
        if resolved_at is None:
            warnings.append(
                f"Строка {row_index} пропущена: не заполнена дата фактического разрешения"
            )
            continue

        items.append(
            {
                "work_date": resolved_at.date(),
                "activity_type": cast(JournalSchemaType, ActivityType.TICKET.value),
                "status": cast(JournalSchemaStatus, ActivityStatus.CLOSED.value),
                "title": _build_excel_import_title(ticket_number=ticket_number),
                "service": service,
                "ticket_number": ticket_number,
                "resolved_at": resolved_at,
            }
        )

    return items, warnings


async def list_entries_for_date(
    session: AsyncSession,
    user_id: UUID,
    day_start: datetime,
    day_end: datetime,
) -> list[ActivityEntry]:
    """Возвращает записи за период по рабочей дате.

    Здесь фильтруем именно по work_date, а не по времени создания записи.
    Это важно, потому что сотрудник может занести запись позже, но она всё равно
    должна попасть в отчёт за исходную рабочую дату.
    """
    repo = JournalRepository(session)
    return await repo.list_for_date_range(
        user_id=user_id,
        date_from=day_start.date(),
        date_to=day_end.date(),
    )


def _combine_work_date_and_time(work_date: date, value: time | None) -> datetime | None:
    """Объединяет дату и время в одно значение для сохранения в базе."""
    if value is None:
        return None
    return datetime.combine(work_date, value, tzinfo=UTC)


def _normalize_optional_text(value: str | None) -> str | None:
    """Нормализует текстовое поле PATCH-запроса, сохраняя явное очищение."""
    if value is None:
        return None
    normalized_value = value.strip()
    return normalized_value or None


def _extract_time_component(value: datetime | None) -> time | None:
    """Возвращает время без таймзоны из datetime, если значение задано."""
    if value is None:
        return None
    return value.timetz().replace(tzinfo=None)


def _resolve_updated_time(
    *,
    field_name: str,
    payload: ActivityEntryUpdateRequest,
    current_value: datetime | None,
) -> time | None:
    """Возвращает итоговое время для PATCH с поддержкой явного очищения."""
    if field_name in payload.model_fields_set:
        return getattr(payload, field_name)
    return _extract_time_component(current_value)


def _resolve_updated_end_date(
    *,
    payload: ActivityEntryUpdateRequest,
    original_work_date: date,
    original_finished_at: datetime | None,
    effective_work_date: date,
) -> date | None:
    """Возвращает итоговую дату закрытия с учётом переноса work_date."""
    if "ended_date" in payload.model_fields_set:
        return payload.ended_date
    if original_finished_at is None:
        return None
    if "work_date" in payload.model_fields_set:
        return effective_work_date + (original_finished_at.date() - original_work_date)
    return original_finished_at.date()


def _parse_bulk_import_text(
    payload: BulkJournalImportRequest,
) -> tuple[list[ParsedTextImportItem], list[str]]:
    """Разбирает текст массового импорта в список нормализованных записей."""
    items: list[ParsedTextImportItem] = []
    warnings: list[str] = []
    current_work_date = payload.default_work_date
    current_status = ActivityStatus.OPEN.value

    for raw_line in payload.text.splitlines():
        line = raw_line.strip()
        if not line:
            continue

        section_match = SECTION_DATE_RE.match(line)
        if section_match:
            section_label = section_match.group("label")
            date_token = section_match.group("date")
            if date_token is not None:
                current_work_date = _parse_import_date(date_token)
            if section_label in {"Выполненные задачи", "Выполнено"}:
                current_status = ActivityStatus.CLOSED.value
            elif section_label in {"Взята в работу", "В работе"}:
                current_status = ActivityStatus.IN_PROGRESS.value
            continue

        entry_status = current_status
        normalized_line = line.lstrip("-•*").strip()
        if normalized_line.startswith("Взята в работу "):
            entry_status = ActivityStatus.IN_PROGRESS.value
            normalized_line = normalized_line[len("Взята в работу ") :].strip()
        elif normalized_line.startswith("Выполненные задачи "):
            entry_status = ActivityStatus.CLOSED.value
            normalized_line = normalized_line[len("Выполненные задачи ") :].strip()
        if not normalized_line:
            continue

        if current_work_date is None:
            raise ValueError(
                "Для массового импорта укажи дату в заголовке секции или передай рабочую дату по умолчанию"
            )

        title, ticket_number, task_url = _parse_import_title_and_links(normalized_line)
        if not title:
            warnings.append(f"Пропущена пустая строка: {raw_line}")
            continue

        activity_type = cast(JournalSchemaType, ActivityType.TASK.value)
        items.append(
            {
                "work_date": current_work_date,
                "activity_type": activity_type,
                "status": cast(JournalSchemaStatus, entry_status),
                "title": title,
                "service": None,
                "ticket_number": ticket_number,
                "task_url": task_url,
            }
        )

    return items, warnings


def _load_excel_workbook(workbook_bytes: bytes):
    """Открывает Excel-файл и превращает ошибки формата в понятную бизнес-ошибку."""
    if not workbook_bytes:
        raise ValueError("Файл Excel пустой")

    try:
        return load_workbook(filename=BytesIO(workbook_bytes), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(
            "Не удалось прочитать Excel-файл. Проверь, что загружен корректный `.xlsx`."
        ) from exc


def _build_excel_header_map(worksheet) -> dict[str, int]:
    """Строит карту индексов нужных столбцов по первой строке Excel-файла."""
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise ValueError("Excel-файл не содержит строк с данными")

    normalized_headers = {
        str(value).strip(): index
        for index, value in enumerate(header_row)
        if value is not None and str(value).strip()
    }

    missing_headers = [
        header_name
        for header_name in _EXCEL_IMPORT_REQUIRED_HEADERS
        if header_name not in normalized_headers
    ]
    if missing_headers:
        raise ValueError(
            "В Excel-файле отсутствуют обязательные столбцы: "
            + ", ".join(missing_headers)
        )

    return {
        internal_name: normalized_headers[source_name]
        for source_name, internal_name in _EXCEL_IMPORT_REQUIRED_HEADERS.items()
    }


def _normalize_excel_ticket_number(value: object) -> str | None:
    """Нормализует номер заявки к формату `SR123456`."""
    normalized_value = _normalize_optional_excel_text(value)
    if not normalized_value:
        return None

    match = TICKET_NUMBER_RE.search(normalized_value)
    if not match:
        return None

    return f"SR{match.group('number')}"


def _normalize_excel_resolved_at(value: object) -> datetime | None:
    """Преобразует значение ячейки Excel в datetime фактического разрешения."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    if isinstance(value, str):
        normalized_value = value.strip()
        if not normalized_value:
            return None
        for date_format in (
            "%d.%m.%Y %H:%M:%S",
            "%d.%m.%Y %H:%M",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
        ):
            try:
                return datetime.strptime(normalized_value, date_format)
            except ValueError:
                continue
    return None


def _normalize_optional_excel_text(value: object) -> str | None:
    """Нормализует текстовую ячейку Excel и убирает пустые значения."""
    if value is None:
        return None
    normalized_value = str(value).strip()
    if not normalized_value:
        return None
    return normalized_value


def _build_excel_import_title(ticket_number: str) -> str:
    """Формирует короткий заголовок записи журнала из номера заявки."""
    return ticket_number[:_MAX_ACTIVITY_TITLE_LENGTH]


def _parse_import_date(value: str) -> date:
    """Парсит дату из заголовка секции вида dd.mm.yy или dd.mm.yyyy."""
    day, month, year = value.split(".")
    parsed_year = int(year)
    if len(year) == 2:
        parsed_year += 2000
    return date(int(parsed_year), int(month), int(day))


def _parse_import_title_and_links(value: str) -> tuple[str, str | None, str | None]:
    """Извлекает заголовок, номер заявки и ссылку из строки импорта."""
    task_url = None
    title = value
    link_match = MARKDOWN_LINK_RE.search(value)
    if link_match:
        title = link_match.group("title").strip()
        task_url = link_match.group("url").strip()

    ticket_number = None
    ticket_match = TICKET_NUMBER_RE.search(title if title else value)
    if ticket_match:
        number = ticket_match.group("number")
        ticket_number = f"SR{number}"
        if title.strip() == number:
            title = f"Задача {number}"
    elif title != value:
        ticket_number = _extract_ticket_number_from_text(value)

    if not ticket_number:
        ticket_number = _extract_ticket_number_from_text(value)

    return title, ticket_number, task_url


def _extract_ticket_number_from_text(value: str) -> str | None:
    """Извлекает номер заявки из произвольного текста."""
    match = TICKET_NUMBER_RE.search(value)
    if match:
        return f"SR{match.group('number')}"
    return None
